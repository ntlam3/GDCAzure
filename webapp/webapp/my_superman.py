#!/usr/bin/env python3
import requests
import openpyxl
import datetime
#import xlwings
def get_month():
    report_month=datetime.date.today().month-1
    if report_month <=0:
        report_month='12'
    elif report_month<10:
        report_month='0'+str(report_month)
    #print(report_year +"-" + report_month)
    else:
        report_month=str(report_month)
    return report_month
def login():
    report_month=get_month()
    if report_month=='12':
        report_year=str(datetime.date.today().year-1)
    else:
        report_year=str(datetime.date.today().year)

    url_open='http://azure.thezabbix.com/superman/login/'
    page = requests.get(url_open)
    cookies = page.cookies

    headers = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'Accept-Language': 'en-US,en;q=0.9',
            'Cache-Control': 'max-age=0',
            'Connection': 'keep-alive',
            # 'Cookie': 'csrftoken=X3k0OoV82SctU6FpuJhsnILLjYJeGsCBJK0R3ijv6mjKOOO03E40yvwYo1Xh4sDi',
            'Origin': 'http://azure.thezabbix.com',
            'Referer': 'http://azure.thezabbix.com/superman/login/',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.5060.134 Safari/537.36 Edg/103.0.1264.77',
    }
    data = {

            'csrfmiddlewaretoken': cookies['csrftoken'],
            'username': 'sdsv.ops',
            'password': '!qusrud00',
            'next': '',
    }
    
    #print(r.text)
    s=requests.session()
    r = s.post(url_open, cookies=cookies, headers=headers, data=data, verify=False)
    #print(response.text)
    file_path="//var//www//webapp//webapp//static//superman"+report_year+report_month+".xlsx"
    print(file_path)
    file_name="/static/superman"+report_year+report_month+".xlsx"
    ##NOTE: The link to download report might be updated every month!!!
    response=s.get('http://azure.thezabbix.com/superman/zabbix_report/download_events?event_date='+report_year+'-'+report_month+'&target=all')
    #response=s.get('http://azure.thezabbix.com/superman/zabbix_report/download_events?event_date=2022-11&target=all')
    f=open(file_path,"wb")
    f.write(response.content)
    f.close()
    return {'file_path':file_path,'file_name':file_name}
def display_info(file_path):
    report_month=get_month()
    report_year=str(datetime.date.today().year)
    wb=openpyxl.load_workbook(file_path)
    wb.sheetnames
    wb.active
    sheet=wb["events"]
    print(sheet.title)
    print(sheet.max_row)
    eu_disaster=0
    eu_high=0
    eu_average=0
    eu_warning=0
    eu_information=0
    crm_disaster=0
    crm_high=0
    crm_average=0
    crm_warning=0
    crm_information=0
    #print(sheet['C'+'2'].value)
    ana_disaster=0
    ana_high=0
    ana_average=0
    ana_warning=0
    ana_information=0
    gdc_disaster=0
    gdc_high=0
    gdc_average=0
    gdc_warning=0
    gdc_information=0
    kcdm_disaster=0
    kcdm_high=0
    kcdm_avg=0
    kcdm_warning=0
    kcdm_information=0
    mena_disaster=0
    mena_high=0
    mena_average=0
    mena_warning=0
    mena_information=0
    grs_disaster=0
    grs_high=0
    grs_average=0
    grs_warning=0
    grs_information=0
    for row in range (2,sheet.max_row+1):
        if sheet['A'+str(row)].value=='eu-promotion':
            if sheet['C'+str(row)].value=='Disaster':
                eu_disaster=eu_disaster+1
            elif sheet['C'+str(row)].value=='High':
                eu_high=eu_high+1
            elif sheet['C'+str(row)].value=='Average':
                eu_average=eu_average+1
            elif sheet['C'+str(row)].value=='Warning':
                eu_warning=eu_warning+1
            elif sheet['C'+str(row)].value=='Information':
                eu_information=eu_information+1
        elif (sheet['A'+str(row)].value=='CDM_EU-WEST_PRD') or (sheet['A'+str(row)].value=='CDM_EU-WEST_PRD, Discovered hosts') or (sheet['A'+str(row)].value=='CDM_US-EAST2_PRD, Discovered hosts') or (sheet['A'+str(row)].value=='CDM_US-EAST2_PRD') or (sheet['A'+str(row)].value=='CIS_CRM_DEVQA') or (sheet['A'+str(row)].value=='CIS_CRM_DEVQA, Discovered hosts') or (sheet['A'+str(row)].value=='CIS_CRM_PRD') or (sheet['A'+str(row)].value=='CIS_CRM_PRD, Discovered hosts') or (sheet['A'+str(row)].value=='CIS_CRM_DEV') or (sheet['A'+str(row)].value=='CIS_CRM_DEV, Discovered hosts'):
            if sheet['C'+str(row)].value=='Disaster':
                crm_disaster=crm_disaster+1
            elif sheet['C'+str(row)].value=='High':
                crm_high=crm_high+1
            elif sheet['C'+str(row)].value=='Average':
                crm_average=crm_average+1
            elif sheet['C'+str(row)].value=='Warning':
                crm_warning=crm_warning+1
            elif sheet['C'+str(row)].value=='Information':
                crm_information=crm_information+1
        elif (sheet['A'+str(row)].value=='CRM_Analytics') or (sheet['A'+str(row)].value=='CRM_Analytics, Discovered hosts'):
            if sheet['C'+str(row)].value=='Disaster':
                ana_disaster=ana_disaster+1
            elif sheet['C'+str(row)].value=='High':
                ana_high=ana_high+1
            elif sheet['C'+str(row)].value=='Average':
                ana_average=ana_average+1
            elif sheet['C'+str(row)].value=='Warning':
                ana_warning=ana_warning+1
            elif sheet['C'+str(row)].value=='Information':
                ana_information=ana_information+1
        elif (sheet['A'+str(row)].value=='GDC') or (sheet['A'+str(row)].value=='GDC, Discovered hosts'):
            if sheet['C'+str(row)].value=='Disaster':
                gdc_disaster=gdc_disaster+1
            elif sheet['C'+str(row)].value=='High':
                gdc_high=gdc_high+1
            elif sheet['C'+str(row)].value=='Average':
                gdc_average=gdc_average+1
            elif sheet['C'+str(row)].value=='Warning':
                gdc_warning=gdc_warning+1
            elif sheet['C'+str(row)].value=='Information':
                gdc_information=gdc_information+1
        elif (sheet['A'+str(row)].value=='NextCDM') or (sheet['A'+str(row)].value=='NextCDM, Discovered hosts'):
            if sheet['C'+str(row)].value=='Disaster':
                kcdm_disaster=kcdm_disaster+1
            elif sheet['C'+str(row)].value=='High':
                kcdm_high=kcdm_high+1
            elif sheet['C'+str(row)].value=='Average':
                kcdm_avg=kcdm_avg+1
            elif sheet['C'+str(row)].value=='Warning':
                kcdm_warning=kcdm_warning+1
            elif sheet['C'+str(row)].value=='Information':
                kcdm_information=kcdm_information+1
        elif (sheet['A'+str(row)].value=='SUB-ME-DWF-PRD') or (sheet['A'+str(row)].value=='SUB-ME-DWF-DEV') or (sheet['A'+str(row)].value=='SUB-ME-DATAENRICH-PRD') or (sheet['A'+str(row)].value=='SUB-ME-IRAN-PRD') or (sheet['A'+str(row)].value=='SUB-ME-IRAN-DEV') or (sheet['A'+str(row)].value=='SUB-ME-TURKEY-PRD') or (sheet['A'+str(row)].value=='SUB-ME-TURKEY-DEV'):
            if sheet['C'+str(row)].value=='Disaster':
                mena_disaster=mena_disaster+1
            elif sheet['C'+str(row)].value=='High':
                mena_high=mena_high+1
            elif sheet['C'+str(row)].value=='Average':
                mena_average=mena_average+1
            elif sheet['C'+str(row)].value=='Warning':
                mena_warning=mena_warning+1
            elif sheet['C'+str(row)].value=='Information':
                mena_information=mena_information+1
        elif (sheet['A'+str(row)].value=='GRS_PRD') or (sheet['A'+str(row)].value=='GRS_DEV')  or (sheet['A'+str(row)].value=='SUB-ME-TDS-PRD') or (sheet['A'+str(row)].value=='SUB-ME-TDS-DEV') or (sheet['A'+str(row)].value=='SUB-ME-AFBI-PRD') or (sheet['A'+str(row)].value=='SUB-ME-AFBI-DEV') or (sheet['A'+str(row)].value=='SUB-ME-MMDM-PRD') or (sheet['A'+str(row)].value=='SUB-ME-MMDM-DEV'):
            if sheet['C'+str(row)].value=='Disaster':
                grs_disaster=grs_disaster+1
            elif sheet['C'+str(row)].value=='High':
                grs_high=grs_high+1
            elif sheet['C'+str(row)].value=='Average':
                grs_average=grs_average+1
            elif sheet['C'+str(row)].value=='Warning':
                grs_warning=grs_warning+1
            elif sheet['C'+str(row)].value=='Information':
                grs_information=grs_information+1
    """
    f=open("C:\Python_projects\webapp\static\\superman"+report_year+report_month+".txt","w")
    f.write("Disaster severity in EU-Promotion: {}\n".format(str(eu_disaster)))
    f.write("High severity in EU-Promotion: {}\n".format(str(eu_high)))
    f.write("Average severity in EU-Promotion: {}\n".format(str(eu_average)))
    f.write("Warning severity in EU-Promotion: {}\n".format(str(eu_warning)))
    f.write("Information severity in EU-Promotion: {}\n".format(str(eu_information)))
    """
    total_eu=eu_disaster+eu_high+eu_average+eu_warning+eu_information
    """
    f.write("Total in EU-Promotion: {}\n".format(str(total_eu)))
    f.write("------------------------------------------\n")
    f.write("Disaster severity in CRM: {}\n".format(str(crm_disaster)))
    f.write("High severity in CRM: {}\n".format(str(crm_high)))
    f.write("Average severity in CRM: {}\n".format(str(crm_average)))
    f.write("Warning severity in CRM: {}\n".format(str(crm_warning)))
    f.write("Information severity in CRM: {}\n".format(str(crm_information)))
    """
    total_crm=crm_disaster+crm_high+crm_average+crm_warning+crm_information
    """
    f.write("Total in CRM: {}\n".format(str(total_crm)))
    f.write("------------------------------------------\n")
    f.write("Disaster severity in CRM_Analyicsc: {}\n".format(str(ana_disaster)))
    f.write("High severity in CRM_Analytics: {}\n".format(str(ana_high)))
    f.write("Average severity in CRM_Analytics: {}\n".format(str(ana_average)))
    f.write("Warning severity in CRM_Analytics: {}\n".format(str(ana_warning)))
    f.write("Information severity in CRM_Analytics: {}\n".format(str(ana_information)))
    """
    total_ana=ana_disaster+ana_high+ana_average+ana_warning+ana_information
    """
    f.write("Total in CRM Analytics: {}\n".format(str(total_ana)))
    f.write("------------------------------------------\n")
    f.write("Disaster severity in GDC India: {}\n".format(str(gdc_disaster)))
    f.write("High severity in GDC India: {}\n".format(str(gdc_high)))
    f.write("Average severity in GDC India: {}\n".format(str(gdc_average)))
    f.write("Warning severity in GDC India: {}\n".format(str(gdc_warning)))
    f.write("Information severity in GDC India: {}\n".format(str(gdc_information)))
    """
    total_gdc=gdc_disaster+gdc_high+gdc_average+gdc_warning+gdc_information
    """
    f.write("Total in GDC India: {}\n".format(str(total_gdc)))
    f.write("------------------------------------------\n")
    f.write("Disaster severity in NextCDM: {}\n".format(str(kcdm_disaster)))
    f.write("High severity in NextCDM: {}\n".format(str(kcdm_high)))
    f.write("Average severity in NextCDM: {}\n".format(str(kcdm_avg)))
    f.write("Warning severity in NextCDM: {}\n".format(str(kcdm_warning)))
    f.write("Information severity in NextCDM: {}\n".format(str(kcdm_information)))
    """
    total_kcdm=kcdm_disaster+kcdm_high+kcdm_avg+kcdm_warning+kcdm_information
    """
    f.write("Total in NextCDM: {}\n".format(str(total_kcdm)))
    f.write("------------------------------------------\n")
    f.write("Disaster severity in MENA: {}\n".format(str(mena_disaster)))
    f.write("High severity in MENA: {}\n".format(str(mena_high)))
    f.write("Average severity in MENA: {}\n".format(str(mena_average)))
    f.write("Warning severity in MENA: {}\n".format(str(mena_warning)))
    f.write("Information severity in MENA: {}\n".format(str(mena_information)))
    """
    total_mena=mena_disaster+mena_warning+mena_high+mena_average+mena_information
    """
    f.write("Total in MENA: {}\n".format(str(total_mena)))
    f.write("------------------------------------------\n")
    f.write("Disaster severity in GRS: {}\n".format(str(grs_disaster)))
    f.write("High severity in GRS: {}\n".format(str(grs_high)))
    f.write("Average severity in GRS: {}\n".format(str(grs_average)))
    f.write("Warning severity in GRS: {}\n".format(str(grs_warning)))
    f.write("Information severity in GRS: {}\n".format(str(grs_information)))
    """
    total_grs=grs_disaster+grs_warning+grs_high+grs_average+grs_information
    """
    f.write("Total in GRS: {}\n".format(str(total_grs)))
    f.write("------------------------------------------\n")
    f.close()
    """
    return {'eu_disaster':eu_disaster,'eu_high':eu_high,'eu_average':eu_average,'eu_warning':eu_warning,'eu_information':eu_information,'total_eu':total_eu,'crm_disaster':crm_disaster,'crm_high':crm_high,'crm_average':crm_average,'crm_warning': crm_warning,'crm_information':crm_information,'total_crm':total_crm,'ana_disaster':ana_disaster,'ana_high':ana_high,'ana_average':ana_average,'ana_warning':ana_warning,'ana_information':ana_information,'total_ana':total_ana,'gdc_disaster':gdc_disaster,'gdc_high':gdc_high,'gdc_average':gdc_average,'gdc_warning':gdc_warning,'gdc_information':gdc_information,'total_gdc':total_gdc,'kcdm_disaster':kcdm_disaster,'kcdm_high':kcdm_high,'kcdm_average':kcdm_avg,'kcdm_warning':kcdm_warning,'kcdm_information':kcdm_information,'total_kcdm':total_kcdm,'mena_disaster':mena_disaster,'mena_high':mena_high,'mena_average':mena_average,'mena_warning':mena_warning,'mena_information':mena_information,'total_mena':total_mena,'grs_disaster':grs_disaster,'grs_high':grs_high,'grs_average':grs_average,'grs_warning':grs_warning,'grs_information':grs_information,'total_grs':total_grs}